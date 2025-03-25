import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

def find_date_start_col(df, sample_row_count=10):
    """
    상위 sample_row_count개의 행을 기준으로 날짜로 변환 가능한 값이 가장 많이 나오는 열을 찾음
    """
    def is_date(x):
        try:
            pd.to_datetime(x)
            return True
        except:
            return False

    max_count = 0
    best_col = None

    for col_idx in range(df.shape[1]):
        col_data = df.iloc[:sample_row_count, col_idx]
        count = sum(col_data.apply(is_date))
        if count > max_count:
            max_count = count
            best_col = col_idx

    if best_col is None:
        raise ValueError("날짜가 포함된 열을 찾을 수 없습니다.")

    return best_col

# 날짜행 탐지 함수
def find_date_row(df, date_start_col=None, min_date_count=1, max_row_check=20):
    """
    DataFrame에서 날짜가 포함된 첫 번째 행(index)을 찾습니다.
    - date_start_col: 날짜가 시작될 것으로 예상되는 열 번호. None이면 자동 탐지
    - min_date_count: 날짜처럼 인식되는 값이 이 개수 이상이면 날짜 행으로 판단
    - max_row_check: 상위 몇 개 행만 검사할지
    """
    if date_start_col is None:
        date_start_col = find_date_start_col(df)

    def is_date(x):
        if isinstance(x, (pd.Timestamp, datetime)):
            return True
        if isinstance(x, str):
            try:
                pd.to_datetime(x)
                return True
            except:
                return False
        return False

    for i in range(min(max_row_check, len(df))):
        row = df.iloc[i, date_start_col:]
        if row.apply(is_date).sum() >= min_date_count:
            return i
    raise ValueError("날짜가 포함된 행을 찾을 수 없습니다.")

# 날짜 열 인덱스 → 날짜 매핑
def get_date_mapping(df, date_row_index, date_start_col=None):
    if date_start_col is None:
        date_start_col = find_date_start_col(df)
        print(date_start_col)

    date_row = df.iloc[date_row_index, date_start_col:]
    dates = pd.to_datetime(date_row, errors='coerce')

    mapping = {}
    for col_idx, date in zip(range(date_start_col, len(df.columns)), dates):
        if pd.notna(date):
            mapping[col_idx] = date
    return mapping

# CTQ별 측정값 추출
def extract_measurement_data(df, date_mapping, date_row_index, search_cols=range(0, 11)):
    results = []

    # 측정POINT 위치 수집
    point_indices = []
    for i in range(date_row_index, len(df)):
        for col in search_cols:
            cell_value = df.iloc[i, col]
            if str(cell_value).strip() == '측정POINT':
                ctq_col = col + 1  # CTQ명은 측정POINT 바로 오른쪽 열에 있음
                ctq_name = df.iloc[i, ctq_col] if ctq_col < df.shape[1] else None
                if pd.notna(ctq_name):
                    point_indices.append((i, ctq_name))
                break  # 찾았으면 다음 행으로

    # CTQ 구간별 측정값 수집
    for idx, (start_idx, ctq_name) in enumerate(point_indices):
        # 다음 CTQ까지 or 마지막 유효 행까지
        if idx + 1 < len(point_indices):
            end_idx = point_indices[idx + 1][0]
        else:
            end_idx = start_idx + 1
            while end_idx < len(df):
                if df.iloc[end_idx, min(date_mapping.keys()):].notna().sum() > 0:
                    end_idx += 1
                else:
                    break

        # start_idx 부터 포함
        for i in range(start_idx, end_idx):
            for col in date_mapping:
                value = df.iloc[i, col]
                if pd.notna(value):
                    results.append({
                        'Date': date_mapping[col],
                        '관리항목명': ctq_name,
                        '측정값': value
                    })

    return pd.DataFrame(results)


# 전체 실행 함수
def process_excel_and_save(input_path, output_path, sheet_name='SUB', save_sheet='변환'):
    df = pd.read_excel(input_path, sheet_name=sheet_name, header=None)
    date_row_idx = find_date_row(df)
    date_map = get_date_mapping(df, date_row_idx)
    df_result = extract_measurement_data(df, date_map, date_row_idx)

    # 정렬
    df_result_sorted = df_result.sort_values(by=["Date", "관리항목명"]).reset_index(drop=True)

    # 저장
    wb = load_workbook(input_path)
    if save_sheet in wb.sheetnames:
        del wb[save_sheet]
    ws_new = wb.create_sheet(save_sheet)
    ws_new.append(["Date", "관리항목명", "측정값"])
    for row in df_result_sorted.itertuples(index=False):
        ws_new.append([row.Date, row.관리항목명, row.측정값])
    wb.save(output_path)
    print(f"✅ 변환 완료: {output_path}")

process_excel_and_save(
    input_path="test00.xlsx",
    output_path="test00_변환완료_함수버전.xlsx"
)