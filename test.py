import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# 📌 파일 경로 (필요시 변경)
excel_path = "test00.xlsx"

# 1. SUB 시트 읽기
df_sub = pd.read_excel(excel_path, sheet_name='SUB', header=None)

# 2. 날짜 행 탐색
def is_date(x):
    if isinstance(x, (pd.Timestamp, datetime)):
        return True
    if isinstance(x, str):
        try:
            pd.to_datetime(x)
            return True
        except:
            return False
    return False

date_row_index = None
for i in range(len(df_sub)):
    row = df_sub.iloc[i, 5:]
    if row.apply(is_date).sum() >= 3:
        date_row_index = i
        break

if date_row_index is None:
    raise ValueError("날짜가 포함된 행을 찾을 수 없습니다.")

# 3. 날짜 열 인덱스 기준 날짜 매핑
date_start_col = 5
date_row = df_sub.iloc[date_row_index, date_start_col:]
dates = pd.to_datetime(date_row, errors='coerce')

date_mapping = {}
for col_idx, date in zip(range(date_start_col, len(df_sub.columns)), dates):
    if pd.notna(date):
        date_mapping[col_idx] = date

# 4. 측정POINT 위치 수집
point_indices = []
for i in range(date_row_index, len(df_sub)):
    if df_sub.iloc[i, 1] == '측정POINT' and pd.notna(df_sub.iloc[i, 2]):
        point_indices.append((i, df_sub.iloc[i, 2]))  # (행번호, CTQ명)

# 5. 측정값 수집 (CTQ 구간별)
results = []

for idx, (start_idx, ctq_name) in enumerate(point_indices):
    # 다음 CTQ까지 또는 유효 데이터 있는 마지막 행까지
    if idx + 1 < len(point_indices):
        end_idx = point_indices[idx + 1][0]
    else:
        # 마지막 CTQ: 값이 있는 마지막 행까지 포함
        end_idx = start_idx + 1
        while end_idx < len(df_sub):
            if df_sub.iloc[end_idx, date_start_col:].notna().sum() > 0:
                end_idx += 1
            else:
                break

    # ✅ CTQ 행(start_idx)부터 포함하여 측정값 추출
    for i in range(start_idx, end_idx):
        for col in date_mapping:
            value = df_sub.iloc[i, col]
            if pd.notna(value):
                results.append({
                    'Date': date_mapping[col],
                    '관리항목명': ctq_name,
                    '측정값': value
                })

# 6. 정렬 및 DataFrame 생성
df_result = pd.DataFrame(results)
df_result_sorted = df_result.sort_values(by=["Date", "관리항목명"]).reset_index(drop=True)

# 7. "변환" 시트로 저장
wb = load_workbook(excel_path)
if "변환" in wb.sheetnames:
    del wb["변환"]
ws_new = wb.create_sheet("변환")
ws_new.append(["Date", "관리항목명", "측정값"])

for row in df_result_sorted.itertuples(index=False):
    ws_new.append([row.Date, row.관리항목명, row.측정값])

# 8. 저장
output_path = "test00_변환완료_최종정리.xlsx"
wb.save(output_path)
print(f"✅ 저장 완료: {output_path}")
